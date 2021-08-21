import os
from openpyxl import load_workbook


def read_xl_data(filename, sheetname):

    parent_dir = os.getcwd()
    filepath = os.path.abspath(parent_dir + "/test_data/excel/" + filename)

    wb = load_workbook(filepath)
    sheet = wb[sheetname]

    print(sheet.max_row)
    print(sheet.max_column)

    data = []

    for row in range(2, sheet.max_row+1):
        row_data = []
        for col in range(1, sheet.max_column+1):
            value = sheet.cell(row=row, column=col).value
            if value is None:
                value = ""
            row_data.append(value)

        data.append(row_data)
    wb.close()
    return data

if __name__ == "__main__":
    print(read_xl_data("login_data.xlsx", "login_fail"))